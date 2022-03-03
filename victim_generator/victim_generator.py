import random
from openpyxl import load_workbook
import csv
import pandas as pd
import numpy as np

n_victims = 10  # number of victims that is generated
min_age = 14  # minimal age of generated victims
max_age = 95  # maximum age of generated victims
random_seed = 10

location_set = set()

wb = load_workbook("victim_generator/namen.xlsx")  # Work Book
ws = wb['Top_eerste_voornamen_NL_2010']  # Work Sheet
column = ws['A']  # Column
names_female = [column[x].value for x in range(len(column))]
column2 = ws['B']  # Column
names_male = [column2[x].value for x in range(len(column2))]

random.seed(random_seed)

victim_data = []
genders = ["Man", "Woman"]
distances = ["short", "middle", "long"]
difficulties = ["easy", "middle", "hard"]
vital_signs = ["low", "middle", "high"]
#0-11 short distance 12-23 middle 24-35 long
locations = [[11,5],[11,6],[12,5],[12,6],[22,3],[22,4],[23,3],[23,4],[22,6],[22,7],[23,6],[23,7],
                [11,14],[11,15],[12,14],[12,15],[22,12],[22,13],[23,12],[23,13],[22,15],[22,16],[23,15],[23,16],
                    [11,23],[11,24],[12,23],[12,24],[22,21],[22,22],[23,21],[23,22],[22,24],[22,25],[23,24],[23,25]]
        
#Generate data for n victims
for n in range(n_victims):
    # generate gender, distance, difficulty, vital_sign and age of victim
    gender = random.choice(genders)
    distance = random.choice(distances)
    difficulty = random.choice(difficulties)
    vital_sign = random.choice(vital_signs)
    age = random.randrange(min_age, max_age)    

    #generate location according to distance
    if distance == "short":
        start_index = 0
        end_index = 11
    if distance == "middle":
        start_index = 12
        end_index = 23
    if distance == "long":
        start_index = 24
        end_index = 35
    while(True):
        location_index = random.randrange(start_index, end_index)
        if(location_index not in location_set):
            location = locations[location_index]
            location_set.add(location_index)
            break
    
    # generate name according to gender
    if gender == "Man":
        name = random.choice(names_male)
    else:
        name = random.choice(names_female)
    victim_data.append([n, name, gender, age, distance, difficulty, location, vital_sign])

victim_data = pd.DataFrame(victim_data, columns=["index", "name", "gender", "age", "distance", "difficulty", "location",
                                                   "vital_sign"])
victim_data.to_csv("victim_generator/victim_data.csv", index=False, sep=";")

print("Written output to 'victim_data.csv'")